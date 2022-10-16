from django.db import models
from django.db.models import F, Value
from django.db.models.functions import Concat
from openpyxl import Workbook


class BaseQuerySet(models.QuerySet):
    excel_sheet_name = "base"
    excel_exclude_fields = ["parent", "id"]
    excel_default_fields = [
        ("xp_account", "Conta XP", F("parent__xp_account_id")),
        ("client_name", "Cliente", F("parent__xp_account__name")),
        ("advisor_id", "Código A", F("parent__xp_account__advisor__advisor_id")),
        (
            "advisor_name",
            "Assessor",
            Concat(
                F("parent__xp_account__advisor__first_name"),
                Value(" "),
                F("parent__xp_account__advisor__last_name"),
            ),
        ),
    ]

    def _excel_fields(self):
        extra_fields = []
        for field in self.model._meta.get_fields():
            if field.name not in self.excel_exclude_fields:
                name = field.name
                verbose_name = field.verbose_name
                reference = F(field.name)
                extra_fields.append((name, verbose_name, reference))
        return extra_fields

    def to_excel(self, wb=None):
        if wb is None:
            wb = Workbook()
            wb.remove_sheet(wb.active)
        ws = wb.create_sheet(self.excel_sheet_name)
        all_fields = self.excel_default_fields + self._excel_fields()
        ws.append([f[1] for f in all_fields])
        qs = self.annotate(**{f[0]: f[2] for f in self.excel_default_fields})
        for row, obj in enumerate(qs, start=2):
            for col, field in enumerate(all_fields, start=1):
                if hasattr(obj, f"get_{field[0]}_display"):
                    value = getattr(obj, f"get_{field[0]}_display")()
                else:
                    value = getattr(obj, field[0])
                ws.cell(row, col).value = str(value if value is not None else "")
        return wb

    def fresh(self):
        return self.filter(parent__fresh=True, parent__xp_account__in_base=True)


class BaseManager(models.Manager):
    queryset = BaseQuerySet

    def get_queryset(self):
        return self.queryset(self.model, using=self._db).annotate(
            xp_account_id=F("parent__xp_account"),
            advisor_id=F("parent__xp_account__advisor__advisor_id"),
        )
