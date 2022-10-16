import io
import time
from base64 import b64encode
from datetime import timedelta

from bull.apps.slack.boletabot import BoletaBot
from bull.apps.xpaccount.exceptions import XPAccountIsInactive
from bull.apps.xpaccount.models.positions import (
    Balance,
    Coe,
    Equity,
    FixedIncome,
    Option,
    OTCOption,
    Positions,
    StructuredProduct,
    StructuredProductLeg,
    Term,
)
from bull.utils import enums, xp_api
from bull.utils.xp_api.exceptions import TreasureNotEnabled
from dateutil.parser import parse
from django.conf import settings
from django.db import models
from django.db.models import F, Q
from django.utils import timezone as tz
from openpyxl import Workbook
from openpyxl.styles import Alignment


class XPAccountQuerySet(models.QuerySet):
    def mesa_rv(self):
        return self.in_base().filter(is_mesa_rv=True)

    def active(self):
        """Pela definição da XP, cliente ativo é o cliente que possui pelo menos
        um dos seguintes:
            1. saldo superior a 100 reais;
            2. saldo inferior a -100 reais;
            3. patrimônio na XP superior a 100 reais;
        """
        return self.in_base().exclude(patrimony__range=(-100, 100))

    def in_base(self):
        return self.filter(in_base=True)

    def least_recently_updated(self):
        return self.in_base().order_by("positions__queried_at")

    def update_positions(self, delay=4):
        for acc in self:
            acc.update_positions()
            print(f"Updated: {acc}")
            time.sleep(delay)

    def birthdays(self, days=5):
        "Clientes que farão aniversário nos próximos `days` dias."
        qs = self.annotate(month=F("birth_date__month"), day=F("birth_date__day"))
        condition = Q()
        for i in range(days + 1):
            date = tz.localdate() + timedelta(days=i)
            condition |= Q(day=date.day, month=date.month)
        return qs.filter(condition).order_by("month", "day")

    # EXCEL
    def _export_mesa_rv(self, wb):
        ws = wb.create_sheet("Mesa RV")
        five_days_ago = tz.localdate() - timedelta(days=5)

        cols = [
            "Nome",
            "Conta XP",
            "Assessor",
            "Referência RV",
            "Saldo Projetado",
            "Alerta - Trend DI",
            "Alerta - TD",
            "Alerta - Saque",
            "Atualizado em",
        ]

        ws.append(cols)

        for row, acc in enumerate(self, start=2):
            ws.cell(row, 1).value = acc.name

            ws.cell(row, 2).value = acc.id
            ws.cell(row, 2).hyperlink = acc.hub_url
            ws.cell(row, 2).style = "Hyperlink"

            ws.cell(row, 3).value = acc.advisor.advisor_id
            ws.cell(row, 3).hyperlink = acc.advisor.slack.url
            ws.cell(row, 3).style = "Hyperlink"

            ws.cell(row, 4).value = acc.mesa_rv_reference
            ws.cell(row, 4).number_format = "R$ #,##0.00"

            ws.cell(row, 5).value = acc.positions.balance.total_projected_balance
            ws.cell(row, 5).number_format = "R$ #,##0.00"

            orders = acc.get_advisor_orders("trend_di", five_days_ago)
            ws.cell(row, 6).value = "Sim" if orders else ""

            try:
                protocols = acc.get_treasure_protocols(five_days_ago)
                ws.cell(row, 7).value = "Sim" if protocols else ""
            except TreasureNotEnabled:
                ws.cell(row, 7).value = "Conta não habilitada para TD"
            except Exception:
                ws.cell(row, 7).value = "Erro desconhecido"

            try:
                withdrawals = acc.notable_withdrawals()
                ws.cell(row, 8).value = "Sim" if withdrawals else ""
            except Exception:
                ws.cell(row, 8).value = "Erro desconhecido"

            date_string = f"{tz.localtime(acc.positions.updated_at):%H:%M %d/%m/%y}"
            ws.cell(row, 9).value = date_string

    def _export_per_asset_type(self, wb):
        ws = wb.create_sheet("Por tipo de ativo")

        cols = [
            ("Conta XP", ""),
            ("Assessor", ""),
            ("Ações", "equity"),
            ("Aluguel", "rental"),
            ("Clubes", "investmentClub"),
            ("COE", "coe"),
            ("Fundos de Investimentos", "investmentFund"),
            ("Fundos Imobiliarios", "realEstate"),
            ("Futuros", "future"),
            ("Opções", "option"),
            ("Ouro", "gold"),
            ("Previdência Privada", "privatePension"),
            ("Produtos Estruturados", "structuredProduct"),
            ("Proventos", "earning"),
            ("Renda Fixa", "fixedIncome"),
            ("Seguro", "insurance"),
            ("Termos", "term"),
            ("Tesouro Direto", "treasure"),
        ]

        ws.cell(1, 1).value = cols[0][0]
        ws.cell(1, 2).value = cols[1][0]

        for col, (name, _) in enumerate(cols[2:], start=2):
            col *= 2
            ws.cell(1, col - 1).value = name
            ws.merge_cells(start_row=1, end_row=1, start_column=col - 1, end_column=col)
            ws.cell(1, col - 1).alignment = Alignment(horizontal="center")

        for row, acc in enumerate(self, start=2):
            ws.cell(row, 1).value = acc.id
            ws.cell(row, 1).hyperlink = acc.hub_url
            ws.cell(row, 1).style = "Hyperlink"

            ws.cell(row, 2).value = acc.advisor.advisor_id
            ws.cell(row, 2).hyperlink = acc.advisor.slack.url
            ws.cell(row, 2).style = "Hyperlink"

            for col, (_, key) in enumerate(cols[2:], start=2):
                pos = acc.positions
                col *= 2

                ws.cell(row, col - 1).value = pos.raw_json[key]["value"]
                ws.cell(row, col - 1).number_format = "R$ #,##0.00"

                ws.cell(row, col).value = pos.raw_json[key].get("percent", 0) * 0.01
                ws.cell(row, col).number_format = "0.0%"

    def _export_structured_product_legs(self, wb):
        ws = wb.create_sheet("Estruturados - Pernas")

        cols = [
            "Conta XP",
            "Assessor",
            "Nome Estrutura",
            "Encerramento",
            "Tipo",
            "Ativo",
            "Fixing",
            "Liquidacao",
            "Tipo Fixing",
            "Qtd Contratada",
            "Qtd Atual",
            "Preco Exercicio",
            "Barreira",
            "Financeiro",
            "Status",
        ]

        ws.append(cols)

        legs = []
        for acc in self:
            for sp in acc.positions.raw_json["structuredProduct"]["details"]:
                for leg in sp["legs"]:
                    leg["acc"] = acc
                    leg["nome_estrutura"] = sp["nomeEstrutura"]
                    leg["encerramento"] = sp["dataEncerramento"]
                    legs.append(leg)

        for row, leg in enumerate(legs, start=2):
            ws.cell(row, 1).value = leg["acc"].id
            ws.cell(row, 1).hyperlink = leg["acc"].hub_url
            ws.cell(row, 1).style = "Hyperlink"

            ws.cell(row, 2).value = leg["acc"].advisor.advisor_id
            ws.cell(row, 2).hyperlink = leg["acc"].advisor.slack.url
            ws.cell(row, 2).style = "Hyperlink"

            ws.cell(row, 3).value = leg["nome_estrutura"]

            ws.cell(row, 4).value = (
                parse(leg["encerramento"]) if leg["encerramento"] else ""
            )
            ws.cell(row, 4).number_format = "d/m/yyyy"

            ws.cell(row, 5).value = leg["tipoOpcaoDescription"]

            ws.cell(row, 6).value = leg["ativo"]

            ws.cell(row, 7).value = (
                parse(leg["dataFixing"]) if leg["dataFixing"] else ""
            )
            ws.cell(row, 7).number_format = "d/m/yyyy"

            ws.cell(row, 8).value = (
                parse(leg["dataLiquidacao"]) if leg["dataLiquidacao"] else ""
            )
            ws.cell(row, 8).number_format = "d/m/yyyy"

            ws.cell(row, 9).value = leg["tipoFixingDescription"]

            ws.cell(row, 10).value = leg["quantidadeContratada"]

            ws.cell(row, 11).value = leg["quantidade"]

            ws.cell(row, 12).value = leg["preco"]
            ws.cell(row, 12).number_format = "R$ #,##0.00"

            ws.cell(row, 13).value = leg["descricao"]

            ws.cell(row, 14).value = leg["financeiro"]
            ws.cell(row, 14).number_format = "R$ #,##0.00"

            ws.cell(row, 15).value = leg["status"]

    def _export_structured_product(self, wb):
        ws = wb.create_sheet("Estruturados")

        cols = [
            "Conta XP",
            "Assessor",
            "Nome Estrutura",
            "Custo",
            "Encerramento",
            "Porcento",
        ]

        ws.append(cols)

        sps = []
        for acc in self:
            for sp in acc.positions.raw_json["structuredProduct"]["details"]:
                sp["acc"] = acc
                sps.append(sp)

        for row, sp in enumerate(sps, start=2):
            ws.cell(row, 1).value = sp["acc"].id
            ws.cell(row, 1).hyperlink = sp["acc"].hub_url
            ws.cell(row, 1).style = "Hyperlink"

            ws.cell(row, 2).value = sp["acc"].advisor.advisor_id
            ws.cell(row, 2).hyperlink = sp["acc"].advisor.slack.url
            ws.cell(row, 2).style = "Hyperlink"

            ws.cell(row, 3).value = sp["nomeEstrutura"]

            ws.cell(row, 4).value = sp["custo"]
            ws.cell(row, 4).number_format = "R$ #,##0.00"

            ws.cell(row, 5).value = (
                parse(sp["dataEncerramento"]) if sp["dataEncerramento"] else ""
            )
            ws.cell(row, 5).number_format = "d/m/yyyy"

            ws.cell(row, 6).value = sp["percent"] * 0.01
            ws.cell(row, 6).number_format = "0.0%"

    def _export_balance(self, wb):
        ws = wb.create_sheet("Saldo")

        cols = [
            # (pretty_name, dict_key)
            ("Conta XP", ""),
            ("Assessor", ""),
            ("Saldo disponível", "availableCash"),
            ("Garantia", "warrantyBmf"),
            ("Resgate de Fundo Pendente", "pendingFunds"),
            ("Resgate de Clube Pendente", "rescuePendingClub"),
            ("Termos a Vencer", "expireTerms"),
            ("Utilização de Conta Margem", "marginAccount"),
            ("Lançamentos Futuros", "futureReleases"),
            ("D+1", "projectedBalanced1"),
            ("D+2", "projectedBalanced2"),
            ("D+3", "projectedBalanced3"),
            ("D > 3", "projectedBalancedOver"),
            ("Saldo Projetado Total", "projectedTotal"),
        ]

        ws.append([i[0] for i in cols])

        for row, acc in enumerate(self, start=2):
            ws.cell(row, 1).value = acc.id
            ws.cell(row, 1).hyperlink = acc.hub_url
            ws.cell(row, 1).style = "Hyperlink"

            ws.cell(row, 2).value = acc.advisor.advisor_id
            ws.cell(row, 2).hyperlink = acc.advisor.slack.url
            ws.cell(row, 2).style = "Hyperlink"

            for col, (_, key) in enumerate(cols[2:], start=3):
                ws.cell(row, col).value = acc.positions.raw_json["balance"][key]
                ws.cell(row, col).number_format = "R$ #,##0.00"

    def _export_stocks(self, wb):
        ws = wb.create_sheet("Ações")

        cols = [
            "Conta XP",
            "Assessor",
            "Cotação",
            "Financeiro",
            "Disponível",
            "Quantidade",
            "Ticker",
            "Custo médio",
            "Performance",
            "Qntd Dia",
            "Qntd Projetada",
            "Qntd Estruturada",
            "Garantia BOV",
            "Garantia BVMF",
            "Custo médio status (?)",
        ]

        ws.append(cols)

        stocks = []
        for acc in self:
            for stock in acc.positions.raw_json["equity"]["details"]:
                stock["acc"] = acc
                stocks.append(stock)

        for row, stock in enumerate(stocks, start=2):
            ws.cell(row, 1).value = stock["acc"].id
            ws.cell(row, 1).hyperlink = stock["acc"].hub_url
            ws.cell(row, 1).style = "Hyperlink"

            ws.cell(row, 2).value = stock["acc"].advisor.advisor_id
            ws.cell(row, 2).hyperlink = stock["acc"].advisor.slack.url
            ws.cell(row, 2).style = "Hyperlink"

            ws.cell(row, 3).value = stock["quote"]
            ws.cell(row, 3).number_format = "R$ #,##0.00"

            ws.cell(row, 4).value = stock["finance"]
            ws.cell(row, 4).number_format = "R$ #,##0.00"

            ws.cell(row, 5).value = stock["avaiable"]  # erro de grafia LOL
            ws.cell(row, 6).value = stock["quantity"]
            ws.cell(row, 7).value = stock["productId"]

            ws.cell(row, 8).value = stock["averageCost"]
            ws.cell(row, 8).number_format = "R$ #,##0.00"

            ws.cell(row, 9).value = stock["performance"] * 0.01
            ws.cell(row, 9).number_format = "0.0%"

            ws.cell(row, 10).value = stock["quantityDay"]
            ws.cell(row, 11).value = stock["quantityProjected"]
            ws.cell(row, 12).value = stock["quantityStructured"]
            ws.cell(row, 13).value = stock["warrantyBOV"]
            ws.cell(row, 14).value = stock["warrantyBVMF"]
            ws.cell(row, 15).value = stock["averageCostStatus"]

    def _export_term(self, wb):
        ws = wb.create_sheet("Termos")

        cols = [
            "Conta XP",
            "Assessor",
            "Ticker",
            "Quantidade",
            "Última cotação",
            "Financeiro",
            "Preço de entrada",
            "Data de Rolagem",
            "Vencimento",
        ]

        ws.append(cols)

        terms = []
        for acc in self:
            for t in acc.positions.raw_json["term"]["details"]:
                t["acc"] = acc
                terms.append(t)

        for row, term in enumerate(terms, start=2):
            ws.cell(row, 1).value = term["acc"].id
            ws.cell(row, 1).hyperlink = term["acc"].hub_url
            ws.cell(row, 1).style = "Hyperlink"

            ws.cell(row, 2).value = term["acc"].advisor.advisor_id
            ws.cell(row, 2).hyperlink = term["acc"].advisor.slack.url
            ws.cell(row, 2).style = "Hyperlink"

            ws.cell(row, 3).value = term["product"]
            ws.cell(row, 4).value = term["quantity"]

            ws.cell(row, 5).value = term["lastQuote"]
            ws.cell(row, 5).number_format = "R$ #,##0.00"

            ws.cell(row, 6).value = term["financial"]
            ws.cell(row, 6).number_format = "R$ #,##0.00"

            ws.cell(row, 7).value = term["entranceFee"]
            ws.cell(row, 7).number_format = "R$ #,##0.00"

            ws.cell(row, 8).value = parse(term["rollingDate"][:19])
            ws.cell(row, 8).number_format = "d/m/yyyy"

            ws.cell(row, 9).value = parse(term["due"][:19])
            ws.cell(row, 9).number_format = "d/m/yyyy"

    def _export_treasure(self, wb):
        ws = wb.create_sheet("Tesouro")

        cols = [
            "Conta XP",
            "Assessor",
            "Subgrupo",
            "Título",
            "Posição",
            "Preço",
            "Quantidade",
            "Disponível",
            "Vencimento",
        ]

        ws.append(cols)

        treasures = []
        for acc in self:
            for subgroup in acc.positions["treasure"]["subGroups"]:
                name = subgroup["name"]
                for item in subgroup["items"]:
                    item["subgroup"] = name
                    item["acc"] = acc
                    treasures.append(item)

        for row, treasure in enumerate(treasures, start=2):
            ws.cell(row, 1).value = treasure["acc"].id
            ws.cell(row, 1).hyperlink = treasure["acc"].hub_url
            ws.cell(row, 1).style = "Hyperlink"

            ws.cell(row, 2).value = treasure["acc"].advisor.advisor_id
            ws.cell(row, 2).hyperlink = treasure["acc"].advisor.slack.url
            ws.cell(row, 2).style = "Hyperlink"

            ws.cell(row, 3).value = treasure["subgroup"]
            ws.cell(row, 4).value = treasure["title"]

            ws.cell(row, 5).value = treasure["position"]
            ws.cell(row, 5).number_format = "R$ #,##0.00"

            ws.cell(row, 6).value = treasure["lastQuote"]
            ws.cell(row, 6).number_format = "R$ #,##0.00"

            ws.cell(row, 7).value = treasure["quantity"]
            ws.cell(row, 8).value = treasure["avaiable"]

            ws.cell(row, 9).value = parse(treasure["due"][:10]).date()
            ws.cell(row, 9).number_format = "d/m/yyyy"

    def new_full_export(self):
        wb = Workbook()
        wb.remove_sheet(wb.active)
        condition = Q(parent__xp_account__in=self) & Q(parent__fresh=True)
        Balance.objects.filter(condition).to_excel(wb)
        Coe.objects.filter(condition).to_excel(wb)
        FixedIncome.objects.filter(condition).to_excel(wb)
        Equity.objects.filter(condition).to_excel(wb)
        Option.objects.filter(condition).to_excel(wb)
        OTCOption.objects.filter(condition).to_excel(wb)
        Term.objects.filter(condition).to_excel(wb)
        StructuredProduct.objects.filter(condition).to_excel(wb)
        StructuredProductLeg.objects.filter(
            structure__parent__xp_account__in=self,
            structure__parent__fresh=True,
        ).to_excel(wb)
        Positions.objects.filter(fresh=True, xp_account__in=self).to_excel(wb)
        return wb

    def full_export(self, send_to):
        wb = Workbook()
        wb.remove_sheet(wb.active)

        self._export_mesa_rv(wb)
        self._export_per_asset_type(wb)
        self._export_structured_product_legs(wb)
        self._export_structured_product(wb)
        self._export_balance(wb)
        self._export_stocks(wb)
        self._export_term(wb)

        with io.BytesIO() as buffer:
            wb.save(buffer)
            wb.close()
            buffer.seek(0)
            BoletaBot.client.files_upload(
                channels=send_to,
                file=buffer.read(),
                initial_comment='Exportação "full" concluída',
                filename=f"full_export_{tz.localtime():%Y-%m-%d_%Hh%M}.xlsx",
                filetype="xlsx",
            )


class XPAccountManager(models.Manager):
    def sync(self):
        """Sincroniza a base de clientes local com a da XP:
        - Atualiza dados;
        - Cria novos clientes;
        - Marca clientes não encontrados com `in_base=False`;

        Retorna uma tupla (inbound, outbound) com a quantidade de clientes.
        """
        customers = xp_api.customers.get_all()
        inbound = 0
        for customer in customers:
            _, created = XPAccount.objects.update_or_create(
                pk=customer.id,
                defaults={
                    **customer.dict(),
                    "in_base": True,
                },
            )
            inbound += created
        outbound = self.exclude(pk__in=[i.id for i in customers]).update(in_base=False)
        return inbound, outbound


class XPAccount(models.Model):
    "Uma conta XP."

    # XP
    id = models.AutoField(
        verbose_name="Conta XP",
        primary_key=True,
    )
    name = models.CharField(
        verbose_name="Nome",
        max_length=128,
        default="CLIENTE DESCONHECIDO",
    )
    cpf = models.CharField(
        verbose_name="CPF",
        max_length=11,
        default="",
    )
    cnpj = models.CharField(
        verbose_name="CNPJ",
        max_length=14,
        default="",
    )
    cellphone = models.CharField(
        verbose_name="Celular",
        max_length=32,
        default="",
    )
    phone = models.CharField(
        verbose_name="Telefone",
        max_length=32,
        default="",
    )
    email = models.EmailField(
        verbose_name="Email",
        default="",
    )
    income = models.DecimalField(
        verbose_name="Renda mensal",
        max_digits=12,
        decimal_places=2,
        default=0,
    )
    occupation = models.CharField(
        verbose_name="Ocupação",
        max_length=128,
        default="",
    )
    state = models.CharField(
        verbose_name="Estado",
        max_length=2,
        choices=enums.State.choices,
        default="",
    )
    birth_date = models.DateField(
        verbose_name="Data de Nascimento",
        null=True,
        blank=True,
    )
    suitability = models.CharField(
        verbose_name="Suitability",
        max_length=12,
        choices=xp_api.enums.Suitability.choices,
        default="",
    )
    suitability_due = models.DateField(
        verbose_name="Vencimento do suitability",
        help_text="Data em que o suitability do cliente ficará desatualizado",
        null=True,
        blank=True,
    )
    investor_qualification = models.CharField(
        verbose_name="Qualificação de Investidor",
        max_length=64,
        choices=xp_api.enums.InvestorQualification.choices,
        default="",
    )
    advisor = models.ForeignKey(
        to=settings.AUTH_USER_MODEL,
        verbose_name="Assessor",
        on_delete=models.PROTECT,
        related_name="clients",
        null=True,
        blank=True,
    )
    amount = models.DecimalField(
        verbose_name="Saldo disponível",
        max_digits=12,
        decimal_places=2,
        default=0,
    )
    patrimony = models.DecimalField(
        verbose_name="Patrimônio na XP",
        help_text="Soma de todos os investimentos na conta",
        max_digits=12,
        decimal_places=2,
        default=0,
    )
    positions = models.OneToOneField(
        to="xpaccount.Positions",
        on_delete=models.PROTECT,
        related_name="linked_to",
        null=True,
        blank=True,
    )

    # Added fields
    created_at = models.DateTimeField(
        verbose_name="Criado em",
        help_text="Data em que o cliente foi adicionado ao banco de dados da Maestro",
        auto_now_add=True,
    )
    updated_at = models.DateTimeField(
        verbose_name="Atualizado em",
        auto_now=True,
    )
    xp_updated_at = models.DateTimeField(
        verbose_name="Atualizado pela XP em",
        help_text="Data/hora em que a XP realizou a atualização do cliente",
        null=True,
    )
    in_base = models.BooleanField(
        verbose_name="Na base de clientes",
        help_text="Se o cliente estiver na base do escritório, esse campos será `True`",
    )
    is_mesa_rv = models.BooleanField(
        verbose_name="Está na Mesa RV",
        help_text="Clientes alinhados com a estrátegia da Mesa RV terão `True` nesse campo",
        default=False,
    )
    mesa_rv_reference = models.DecimalField(
        verbose_name="Referência Mesa RV",
        max_digits=12,
        decimal_places=2,
        null=True,
        blank=True,
    )
    hubspot_id = models.IntegerField(
        verbose_name="ID HubSpot",
        null=True,
        blank=True,
    )

    objects = XPAccountManager.from_queryset(XPAccountQuerySet)()

    def __str__(self):
        return str(self.pk)

    @classmethod
    def __get_validators__(cls):
        "Permite que a classe seja usada com o Pydantic"
        yield lambda *args, **kwargs: int(*args, **kwargs)
        # O yield abaixo garante que se v == 0, o valor retornado será None
        # já que a XP de vez em quando coloca 0 nos relatórios de receita
        # para indicar que a receita não está associada a nenhum assessor.
        # Quando v != 0, retornará um erro, como deveria, caso v não seja
        # o número de uma conta XP válida.
        yield lambda v: cls.objects.get(pk=v) if v else None

    class Meta:
        verbose_name = "Conta XP"
        verbose_name_plural = "Contas XP"
        ordering = ["id"]
        get_latest_by = "positions__updated"

    @property
    def is_active(self):
        if self.patrimony is None:
            return False
        return self.patrimony >= 100 or self.patrimony <= -100

    def get_advisor_orders(self, fund, start=None, end=None):
        fund_ids = {
            "trend_di": "0fad93c4-ce0d-4d39-9152-4bafd339fac1",
        }
        fund = fund_ids.get(fund, fund)
        return xp_api.advisor_orders.get(self.id, fund, start, end).data

    def get_balance(self):
        return xp_api.balance.get(self.id)

    def get_customer_info(self):
        return xp_api.customer_info.get(self.id).customer_info

    def get_treasure_protocols(self, start=None, end=None):
        return xp_api.treasure_protocols.get(self.id, start, end).__root__

    def get_withdrawals(self):
        return xp_api.withdrawals.get(self.id).historical_withdrawal

    def get_positions(self):
        return xp_api.positions.get(self.id)

    @property
    def hub_url(self):
        "URL do cliente no HUB XP."
        host = "https://hub.xpi.com.br"
        account = b64encode(str(self.id).encode("ascii")).decode("utf-8")
        return f"{host}/rede/#/customers/{account}/consolidated-position"

    def age(self):
        "Calcula a idade do cliente."
        now = tz.localdate()
        age = now.year - self.birth_date.year
        if now.month < self.birth_date.month and now.day < self.birth_date.day:
            age -= 1
        return age

    def update_positions(self):
        """Consulta a API da XP para pegar os dados mais atualizados da carteira
        do cliente e salva no banco de dados."""
        if not self.in_base:
            raise XPAccountIsInactive
        self.positions = Positions.objects.new(self.id)
        self.save()

    def notable_withdrawals(self):
        return [w for w in self.get_withdrawals() if w.is_notable()]
